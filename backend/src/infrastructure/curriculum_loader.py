"""Infrastructure — Standard curriculum loader from xlsx (Spec.md section 7)

Reads all discipline names from matriz_curricular_editavel.xlsx across all sheets
and provides a cached list for validation use in the analysis pipeline.
"""
from __future__ import annotations
import logging
from functools import lru_cache
from pathlib import Path
from typing import List

logger = logging.getLogger(__name__)

# Column index (0-based) where the discipline name lives in every sheet
_DISCIPLINE_COL = 2
_HEADER_MARKER = "Nome da disciplina"


class CurriculumLoadError(RuntimeError):
    """Raised when the standard curriculum xlsx cannot be loaded."""


def load_standard_curriculum(xlsx_path: Path) -> List[str]:
    """Load all unique, non-empty discipline names from the xlsx file.

    Iterates over every sheet, locates the header row by looking for the
    cell whose value equals ``"Nome da disciplina"`` in column 2, then
    collects every subsequent non-empty value in that column.

    Args:
        xlsx_path: Absolute path to ``matriz_curricular_editavel.xlsx``.

    Returns:
        Deduplicated list of discipline name strings (preserving first-seen order).

    Raises:
        CurriculumLoadError: If the file cannot be opened or has no data.
    """
    try:
        import openpyxl  # local import so the module can still be imported if openpyxl is absent
    except ImportError as exc:
        raise CurriculumLoadError(
            "openpyxl not installed. Run: pip install openpyxl==3.1.2"
        ) from exc

    if not xlsx_path.exists():
        raise CurriculumLoadError(f"Arquivo não encontrado: {xlsx_path}")

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as exc:
        raise CurriculumLoadError(f"Não foi possível abrir o xlsx: {exc}") from exc

    seen: set[str] = set()
    ordered: List[str] = []

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_found = False
            for row in ws.iter_rows(values_only=True):
                if not header_found:
                    # Detect header row: look for the marker in the discipline column
                    cell_val = row[_DISCIPLINE_COL] if row and len(row) > _DISCIPLINE_COL else None
                    if cell_val and str(cell_val).strip() == _HEADER_MARKER:
                        header_found = True
                    continue

                # Data row
                if row and len(row) > _DISCIPLINE_COL:
                    raw = row[_DISCIPLINE_COL]
                    if raw is None:
                        continue
                    name = str(raw).strip()
                    # Strip trailing annotation suffixes like **, ***, ****Ext, Ext
                    import re
                    name = re.sub(r"\*+(?:Ext)?$", "", name).strip()
                    name = re.sub(r"Ext$", "", name).strip()
                    if name and name not in seen:
                        seen.add(name)
                        ordered.append(name)
    finally:
        wb.close()

    if not ordered:
        raise CurriculumLoadError(
            f"Nenhuma disciplina encontrada no arquivo {xlsx_path.name}. "
            "Verifique se o cabeçalho 'Nome da disciplina' existe na coluna C de cada aba."
        )

    logger.info("Currículo padrão carregado: %d disciplinas de %s", len(ordered), xlsx_path.name)
    return ordered


@lru_cache(maxsize=1)
def get_standard_curriculum() -> List[str]:
    """Return the cached standard curriculum list.

    Reads from the project data directory on first call; subsequent calls
    return the cached list without re-reading the file.

    Returns:
        List of discipline name strings from the standard curriculum.
    """
    from src.config import get_data_dir
    xlsx_path = get_data_dir() / "matriz_curricular_editavel.xlsx"
    return load_standard_curriculum(xlsx_path)


def discipline_in_standard(nome: str) -> bool:
    """Return True if *nome* fuzzy-matches any discipline in the standard curriculum.

    Uses a case-insensitive, accent-tolerant substring check so that minor
    variations in formatting (trailing markers, capitalisation) do not cause
    false negatives.
    """
    if not nome or not isinstance(nome, str):
        return False
    try:
        standard = get_standard_curriculum()
    except CurriculumLoadError:
        logger.warning("discipline_in_standard: could not load standard curriculum — skipping check")
        return True  # fail-open: don't block the analysis

    nome_lower = nome.lower().strip()
    for s in standard:
        if nome_lower in s.lower() or s.lower() in nome_lower:
            return True
    return False

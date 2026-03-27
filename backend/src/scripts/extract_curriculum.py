"""Script para extrair disciplinas de uma matriz curricular (XLSX) para JSON (Spec.md).

Uso:
  python src/scripts/extract_curriculum.py [caminho_xlsx] [nome_curso]

Exemplo:
  python src/scripts/extract_curriculum.py data/matriz_curricular_editavel.xlsx Administração
"""
import sys
import json
import logging
from pathlib import Path

# Ajustando o sys.path para importações do src caso seja rodado diretamente
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from src.config import get_data_dir

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")


def sanitize_filename(name: str) -> str:
    """Remove caracteres inválidos e acentos para nomes de pastas e arquivos."""
    import re
    import unicodedata
    s = unicodedata.normalize('NFKD', name).encode('ASCII', 'ignore').decode('utf-8')
    s = re.sub(r'[^a-zA-Z0-9_\-]', '_', s)
    return s.strip('_')


def clean_string(val) -> str:
    """Limpa a string removendo espaços e asteriscos de notação extra."""
    if val is None:
        return ""
    s = str(val).strip()
    import re
    s = re.sub(r"\*+(?:Ext)?$", "", s).strip()
    s = re.sub(r"Ext$", "", s).strip()
    return s


def parse_xlsx_to_json(xlsx_path: Path, curso_nome: str):
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl não instalado. Execute: pip install openpyxl")
        return

    if not xlsx_path.exists():
        logger.error(f"Arquivo não encontrado: {xlsx_path}")
        return

    # Diretório destino: data/CursosUFSM/{curso_nome}/
    dest_dir = get_data_dir() / "CursosUFSM" / sanitize_filename(curso_nome)
    dest_dir.mkdir(parents=True, exist_ok=True)

    logger.info(f"Lendo {xlsx_path.name} para o curso de {curso_nome}...")
    
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    
    # Mapeamento de cabeçalhos esperados para nossa lógica interna
    expected_headers = {
        "código": ["código", "codigo"],
        "nome_disciplina": ["nome da disciplina", "disciplina"],
        "semestre": ["semestre", "fase", "período"],
        "tipo": ["tipo", "caráter", "obr/opt"],
        "t_p_pext": ["t-p-pext", "ch t-p-e"],
        "ch_total": ["ch total", "carga horária"],
        "nucleo_grupo": ["núcleo/grupo", "núcleo", "grupo"],
        "pre_requisito": ["pré-requisito", "pré-requisitos", "requisitos"]
    }

    count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        header_map = {}
        header_row_idx = -1
        
        # Encontra a linha de cabeçalho
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            row_strs = [str(c).strip().lower() for c in row if c is not None]
            
            # Se a linha contém pelo menos "código" e "nome", assumimos que é o cabeçalho
            if any(h in row_strs for h in expected_headers["código"]) and any(h in row_strs for h in expected_headers["nome_disciplina"]):
                header_row_idx = row_idx
                # Mapeia o índice da coluna para o nome canônico
                for col_idx, cell in enumerate(row):
                    if cell is None: continue
                    val = str(cell).strip().lower()
                    for canon_name, aliases in expected_headers.items():
                        if val in aliases:
                            header_map[canon_name] = col_idx
                            break
                break
                
        if not header_map or "código" not in header_map or "nome_disciplina" not in header_map:
            logger.debug(f"Aba '{sheet_name}' ignorada (cabeçalhos obrigatórios não encontrados).")
            continue
            
        # Extrai os dados a partir da linha logo após o cabeçalho
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx <= header_row_idx:
                continue
                
            codigo_raw = row[header_map["código"]] if header_map.get("código") is not None and len(row) > header_map["código"] else None
            nome_raw = row[header_map["nome_disciplina"]] if header_map.get("nome_disciplina") is not None and len(row) > header_map["nome_disciplina"] else None
            
            codigo = clean_string(codigo_raw)
            nome = clean_string(nome_raw)
            
            if not codigo or not nome:
                continue # Ignora linhas vazias ou de totais
                
            # Extração dos outros campos
            semestre_raw = row[header_map["semestre"]] if "semestre" in header_map and len(row) > header_map["semestre"] else None
            sem_val = (semestre_raw if isinstance(semestre_raw, int) or str(semestre_raw).isdigit() else None)
            
            tipo = clean_string(row[header_map["tipo"]] if "tipo" in header_map and len(row) > header_map["tipo"] else None)
            
            ch_raw = row[header_map["ch_total"]] if "ch_total" in header_map and len(row) > header_map["ch_total"] else None
            ch_total = int(ch_raw) if ch_raw and str(ch_raw).isdigit() else 0
            
            nucleo = clean_string(row[header_map["nucleo_grupo"]] if "nucleo_grupo" in header_map and len(row) > header_map["nucleo_grupo"] else None)
            t_p_pext = clean_string(row[header_map["t_p_pext"]] if "t_p_pext" in header_map and len(row) > header_map["t_p_pext"] else None)
            
            # Pré-requisitos
            prereq_raw = row[header_map["pre_requisito"]] if "pre_requisito" in header_map and len(row) > header_map["pre_requisito"] else None
            prereq_list = []
            if prereq_raw:
                # Pode usar vírgula ou " e " para separar
                import re
                parts = re.split(r'[,;]|\be\b', str(prereq_raw))
                prereq_list = [p.strip() for p in parts if p.strip()]

            # Monta o dicionário
            data = {
                "codigo": codigo,
                "nome": nome,
                "semestre": int(sem_val) if sem_val is not None else None,
                "tipo": tipo.upper() if tipo else "OBR", # default fallback requested by user as OBR/ELE etc
                "ch_total": ch_total,
                "nucleo_grupo": nucleo,
                "programa": "", # Campo para preenchimento da coordenação
                "pre_requisito": prereq_list,
                "t_p_pext": t_p_pext
            }
            
            # Salva o JSON
            file_name = f"{sanitize_filename(codigo)}.json"
            out_file = dest_dir / file_name
            with open(out_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            count += 1

    wb.close()
    logger.info(f"Concluído: {count} disciplinas extraídas e salvas na pasta {dest_dir}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python extract_curriculum.py [planilha.xlsx] [Nome_do_Curso]")
        sys.exit(1)
        
    p_xlsx = Path(sys.argv[1])
    c_nome = sys.argv[2]
    
    parse_xlsx_to_json(p_xlsx, c_nome)
